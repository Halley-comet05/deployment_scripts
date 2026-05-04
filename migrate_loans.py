# migrate_loans.py  
# Run with: bench --site your-site.local execute migrate_loans.migrate_all  
  
import frappe  
import openpyxl  
from frappe.utils import flt, getdate, rounded
from lending.loan_management.doctype.loan_demand.loan_demand import create_loan_demand  
from lending.loan_management.doctype.loan_repayment_schedule.utils import set_demand
  
from datetime import date, datetime

EXCEL_PATH = "/workspace/development/frappe-bench/Import_Loans_Demo.xlsx"  
  
def migrate_all():  
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)  
    ws = wb.active  
  
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]  
  
    results = {"success": [], "failed": []}  
  
    for row in ws.iter_rows(min_row=2, values_only=True):  
        data = dict(zip(headers, row))  
  
        if not data.get("ID"):  
            continue  
  
        try:  
            migrate_loan(data)  
            results["success"].append(data["ID"])  
            frappe.db.commit()  
        except Exception as e:  
            frappe.db.rollback()  
            results["failed"].append({  
                "loan": data["ID"],  
                "error": str(e)[:200]  
            })  
            print(f"  Error migrating {data['ID']}: {str(e)[:200]}")  
  
    print(f"\nMigration complete.")  
    print(f"  Success: {len(results['success'])}")  
    print(f"  Failed:  {len(results['failed'])}")  
  
    if results["failed"]:  
        print("\nFailed loans:")  
        for f in results["failed"]:  
            print(f"  {f['loan']}: {f['error']}")  
  
    return results  
  
  
def migrate_loan(data):  
    loan_name = data["ID"]  
    payments_made = int(data.get("payments_made") or 0)  
    loan_amount = rounded(data["Loan Amount"], 2)  
    rate_of_interest = rounded(data["Rate of Interest per year"], 2)  
    repayment_periods = int(data["Tenure"])  
    repayment_start_date = to_date(data["repayment Start date"])
    repayment_frequency = data["Repayment Schedule Type"]  
    posting_date = to_date(data["Posting Date"])
    fees = rounded(data["Fees"], 2)
  
    if frappe.db.exists("Loan", loan_name):  
        print(f"  Skipping {loan_name} — already exists")  
        return  
  
    # --- 1. Create Loan ---  
    loan = frappe.new_doc("Loan")  
    loan.name = loan_name  
    loan.applicant_type = data["Applicant Type"]  
    loan.applicant = data["Applicant"]  
    loan.company = data["Company"]  
    loan.posting_date = posting_date  
    loan.loan_product = data["Loan Product"]  
    loan.is_term_loan = 1
    loan.rate_of_interest = rate_of_interest  
    loan.loan_amount = loan_amount  
    loan.repayment_start_date = repayment_start_date  
    loan.repayment_periods = repayment_periods  
    loan.repayment_frequency = repayment_frequency  
    loan.repayment_method = "Repay Over Number of Periods"
    loan.is_flat_rate_interest = 1  
    loan.status = "Sanctioned"
    loan.custom_account_number = data.get("Account Number")
    loan.custom_mode_settlement = data.get("Mode of Settlement")
      
    # Set accounts from loan product  
    loan_product_details = frappe.db.get_value("Loan Product", data["Loan Product"], [  
        "loan_account", "payment_account", "interest_income_account",   
        "disbursement_account", "penalty_income_account"  
    ], as_dict=1)  
      
    if loan_product_details:  
        loan.loan_account = loan_product_details.loan_account  
        loan.payment_account = loan_product_details.payment_account  
        loan.interest_income_account = loan_product_details.interest_income_account  
        loan.disbursement_account = loan_product_details.disbursement_account  
        loan.penalty_income_account = loan_product_details.penalty_income_account  
      
    loan.flags.ignore_mandatory = True  
    loan.insert(ignore_permissions=True)  
    loan.submit()  
    frappe.db.commit()  
    actual_loan_name = loan.name  
    print(f"  Loan created as: {actual_loan_name}")  
  
    # --- 2. Create Loan Disbursement ---  
    disbursement = frappe.new_doc("Loan Disbursement")  
    disbursement.against_loan = actual_loan_name  
    disbursement.applicant_type = data["Applicant Type"]  
    disbursement.applicant = data["Applicant"]  
    disbursement.company = data["Company"]  
    disbursement.posting_date = posting_date  
    disbursement.disbursement_date = posting_date
    disbursement.disbursed_amount = loan_amount  
    disbursement.custom_upfront_fees = fees
    disbursement.repayment_start_date = repayment_start_date  
    disbursement.repayment_periods = repayment_periods  
    disbursement.repayment_frequency = repayment_frequency  
    disbursement.rate_of_interest = rate_of_interest  
    disbursement.repayment_method = "Repay Over Number of Periods"
    disbursement.is_flat_rate_interest = 1  
    disbursement.principal_share_percentage = 100  
    disbursement.interest_share_percentage = 100  
    disbursement.tenure = repayment_periods  
      
    disbursement.flags.ignore_mandatory = True  
    disbursement.insert(ignore_permissions=True)  
    frappe.db.commit()  
  
    # --- 3. Generate repayment schedule ---  
    disbursement.make_update_draft_schedule()  
    frappe.db.commit()  
  
    disbursement.reload()  
    disbursement.submit()  
    frappe.db.commit()  
    print(f"  Disbursement submitted: {disbursement.name}")

    schedule_name = frappe.db.get_value("Loan Repayment Schedule", {"loan_disbursement": disbursement.name}, "name")
    if schedule_name and payments_made > 0:
        schedule = frappe.get_doc("Loan Repayment Schedule", schedule_name)
        monthly_repayment_amount = rounded(schedule.repayment_schedule[0].total_payment, 2) if schedule.repayment_schedule else 0
        schedule.reload()

        for i, row in enumerate(schedule.repayment_schedule):  
            if i >= payments_made:  
                break  

            if row.principal_amount:  
                create_loan_demand(  
                    loan=actual_loan_name,  
                    demand_date=row.payment_date,  
                    demand_type="EMI",  
                    demand_subtype="Principal",  
                    amount=rounded(row.principal_amount, 2),  
                    loan_repayment_schedule=schedule_name,  
                    loan_disbursement=disbursement.name,  
                    repayment_schedule_detail=row.name,  
                    posting_date=row.payment_date,  
                    paid_amount=rounded(row.principal_amount, 2)  
                )  

            if row.interest_amount:  
                create_loan_demand(  
                    loan=actual_loan_name,  
                    demand_date=row.payment_date,  
                    demand_type="EMI",  
                    demand_subtype="Interest",  
                    amount=rounded(row.interest_amount, 2),  
                    loan_repayment_schedule=schedule_name,  
                    loan_disbursement=disbursement.name,  
                    repayment_schedule_detail=row.name,  
                    posting_date=row.payment_date,  
                    paid_amount=rounded(row.interest_amount, 2)  
                )  

            frappe.db.set_value("Repayment Schedule", row.name, "demand_generated", 1)

        frappe.db.commit()  

        total_amount_paid = rounded(data.get("Total Amount Paid") or 0, 2)  
        total_principal_paid = rounded(data.get("Total Principal Paid") or 0, 2)  

        frappe.db.set_value("Loan Repayment Schedule", schedule_name, {  
            "total_installments_paid": payments_made,  
            "total_installments_raised": payments_made,  
            "monthly_repayment_amount": monthly_repayment_amount,  
        })  
        frappe.db.set_value("Loan", actual_loan_name, {  
            "total_amount_paid": total_amount_paid,  
            "total_principal_paid": total_principal_paid,
            "monthly_repayment_amount": monthly_repayment_amount,
            "is_term_loan": 1,
        })

    else:
        if schedule_name:
            schedule = frappe.get_doc("Loan Repayment Schedule", schedule_name)
            monthly_repayment_amount = rounded(schedule.repayment_schedule[0].total_payment, 2) if schedule.repayment_schedule else 0
            frappe.db.set_value("Loan Repayment Schedule", schedule_name, {  
                "monthly_repayment_amount": monthly_repayment_amount,  
            })  
            frappe.db.set_value("Loan", actual_loan_name, {  
                "monthly_repayment_amount": monthly_repayment_amount,
                "is_term_loan": 1,
            })

    if schedule_name:
        frappe.db.set_value("Loan Disbursement", disbursement.name, "monthly_repayment_amount", monthly_repayment_amount)
        frappe.db.commit()
  
    print(f"  Migrated {loan_name} as {actual_loan_name} ({payments_made} payments marked paid)")

def to_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value          # already a proper date, use as-is
    if value:
        return getdate(str(value).split(" ")[0])  # fallback for string values
    return None
