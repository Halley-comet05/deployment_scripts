import math
import pandas as pd
from openpyxl import load_workbook

# ── CONFIG ──────────────────────────────────────────────────────────────────
SOURCE_FILE    = "Loans EDITED.xlsx"
SOURCE_SHEET   = "Loan"
TEMPLATE_FILE  = "Loans.xlsx"
OUTPUT_FILE    = "Loans_import.xlsx"
COMPANY        = "SSEMPC"
REPAYMENT_FREQ = "Bi-Monthly (15th & 28th)"
# ────────────────────────────────────────────────────────────────────────────

df = pd.read_excel(SOURCE_FILE, sheet_name=SOURCE_SHEET, dtype=str)
df = df.fillna("")

wb = load_workbook(TEMPLATE_FILE)
ws = wb.active

headers = [cell.value for cell in ws[1]]

def clean_date(val):
    val = str(val).strip()
    if " " in val:
        val = val.split(" ")[0]
    return val if val != "nan" else ""

def round_down_payments(val):
    val = val.strip()
    if not val:
        return ""
    try:
        return math.floor(float(val))
    except ValueError:
        return ""

def round_2(val):
    val = val.strip()
    if not val:
        return ""
    try:
        return round(float(val), 2)
    except ValueError:
        return ""

for i, row in df.iterrows():
    excel_row = i + 2

    interest_x_tenure = row.get("Interest x Tenure", "").strip()
    try:
        rate = round(float(interest_x_tenure), 2)
    except ValueError:
        rate = ""

    data = {
        "ID": "",
        "Applicant Type": "Customer",
        "Applicant": row.get("Applicant", "").strip(),
        "Company": COMPANY,
        "Posting Date": clean_date(row.get("Posting Date", "")),
        "Loan Product": row.get("Loan Product", "").strip(),
        "Rate of Interest per year": rate,
        "Loan Amount": row.get("principal", "").strip(),
        "repayment Start date": clean_date(row.get("Repayment Start Date", "")),
        "Tenure": row.get("term", "").strip(),
        "Repayment Schedule Type": REPAYMENT_FREQ,
        "Enable Flat-Rate Interest": 1,
        "Disbursement Date": clean_date(row.get("Disbursement Date", "")),
        "Disbursed Amount": row.get("Disbursed Amount", "").strip(),
        "payments_made": round_down_payments(row.get("Payments Made", "")),
        "Total Amount Paid": round_2(row.get("Total Amount Paid", "")),
        "Total Principal Paid": round_2(row.get("Total Principal Paid", "")),
    }

    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=excel_row, column=col_idx, value=data.get(header, ""))

wb.save(OUTPUT_FILE)
print(f"Done! Saved to {OUTPUT_FILE}")
